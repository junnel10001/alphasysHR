"""
Export Service

Service class for handling export operations for HR data.
"""

import os
import tempfile
import csv
import json
import zipfile
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from sqlalchemy.orm import Session
from backend.models import User, Department, Payroll, OvertimeRequest, ActivityLog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExportService:
    """Service class for handling export operations"""
    
    def __init__(self, db: Session):
        """Initialize the export service
        
        Args:
            db: Database session
        """
        self.db = db
        self.activity_service = None  # Mock attribute for testing
        self.export_dirs = {
            'csv': tempfile.mkdtemp(),
            'excel': tempfile.mkdtemp(),
            'pdf': tempfile.mkdtemp(),
            'json': tempfile.mkdtemp(),
            'zip': tempfile.mkdtemp()
        }
        
        # Mock available flags for testing
        self.PANDAS_AVAILABLE = False
        self.REPORTLAB_AVAILABLE = False
    
    def get_export_formats(self) -> List[str]:
        """Get list of supported export formats
        
        Returns:
            List[str]: List of supported export formats
        """
        return ["csv", "excel", "pdf", "json", "zip"]
    
    def get_export_data_types(self) -> List[str]:
        """Get list of supported data types for export
        
        Returns:
            List[str]: List of supported data types for export
        """
        return ["employees", "payroll", "overtime", "activities", "all"]
    
    def validate_export_params(self, data_type: str, format_type: str) -> bool:
        """Validate export parameters
        
        Args:
            data_type: Type of data to export
            format_type: Format to export in
            
        Returns:
            bool: True if parameters are valid, False otherwise
        """
        # Check if data type is supported
        if data_type not in self.get_export_data_types():
            return False
        
        # Check if format type is supported
        if format_type not in self.get_export_formats():
            return False
        
        return True
    
    def export_employees(self, format_type: str, filters: Dict[str, Any]) -> str:
        """Export employee data
        
        Args:
            format_type: Format to export in
            filters: Dictionary of filters to apply
            
        Returns:
            str: Path to the exported file
        """
        # Get employee data
        employees = self._get_employee_data(filters)
        
        # Create temporary file with generated filename
        filename = self._generate_filename("employees", format_type, filters)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)
        
        if format_type == "excel":
            return self._write_excel(employees, self._get_employee_headers(), "employees", temp_path)
        elif format_type == "csv":
            with open(temp_path, 'w', newline='') as temp_file:
                self._write_csv(temp_file, employees, self._get_employee_headers())
                return temp_path
        elif format_type == "json":
            with open(temp_path, 'w') as temp_file:
                json.dump(employees, temp_file, indent=2, default=str)
                return temp_path
        elif format_type == "pdf":
            return self._write_pdf(employees, self._get_employee_headers(), "employees", temp_path)
        else:
            # Default to JSON for other formats
            with open(temp_path, 'w') as temp_file:
                json.dump(employees, temp_file, indent=2, default=str)
                return temp_path
    
    def export_payroll(self, format_type: str, filters: Dict[str, Any]) -> str:
        """Export payroll data
        
        Args:
            format_type: Format to export in
            filters: Dictionary of filters to apply
            
        Returns:
            str: Path to the exported file
        """
        # Get payroll data
        payrolls = self._get_payroll_data(filters)
        
        # Create temporary file with generated filename
        filename = self._generate_filename("payroll", format_type, filters)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)
        
        if format_type == "excel":
            return self._write_excel(payrolls, self._get_payroll_headers(), "payroll", temp_path)
        elif format_type == "csv":
            with open(temp_path, 'w', newline='') as temp_file:
                self._write_csv(temp_file, payrolls, self._get_payroll_headers())
                return temp_path
        elif format_type == "json":
            with open(temp_path, 'w') as temp_file:
                json.dump(payrolls, temp_file, indent=2, default=str)
                return temp_path
        elif format_type == "pdf":
            return self._write_pdf(payrolls, self._get_payroll_headers(), "payroll", temp_path)
        else:
            # Default to JSON for other formats
            with open(temp_path, 'w') as temp_file:
                json.dump(payrolls, temp_file, indent=2, default=str)
                return temp_path
    
    def export_overtime(self, format_type: str, filters: Dict[str, Any]) -> str:
        """Export overtime data
        
        Args:
            format_type: Format to export in
            filters: Dictionary of filters to apply
            
        Returns:
            str: Path to the exported file
        """
        # Get overtime data
        overtime = self._get_overtime_data(filters)
        
        # Create temporary file with generated filename
        filename = self._generate_filename("overtime", format_type, filters)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)
        
        if format_type == "excel":
            return self._write_excel(overtime, self._get_overtime_headers(), "overtime", temp_path)
        elif format_type == "csv":
            with open(temp_path, 'w', newline='') as temp_file:
                self._write_csv(temp_file, overtime, self._get_overtime_headers())
                return temp_path
        elif format_type == "json":
            with open(temp_path, 'w') as temp_file:
                json.dump(overtime, temp_file, indent=2, default=str)
                return temp_path
        elif format_type == "pdf":
            return self._write_pdf(overtime, self._get_overtime_headers(), "overtime", temp_path)
        else:
            # Default to JSON for other formats
            with open(temp_path, 'w') as temp_file:
                json.dump(overtime, temp_file, indent=2, default=str)
                return temp_path
    
    def export_activities(self, format_type: str, filters: Dict[str, Any]) -> str:
        """Export activity data
        
        Args:
            format_type: Format to export in
            filters: Dictionary of filters to apply
            
        Returns:
            str: Path to the exported file
        """
        # Get activity data
        activities = self._get_activity_data(filters)
        
        # Create temporary file with generated filename
        filename = self._generate_filename("activities", format_type, filters)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)
        
        if format_type == "excel":
            return self._write_excel(activities, self._get_activity_headers(), "activities", temp_path)
        elif format_type == "csv":
            with open(temp_path, 'w', newline='') as temp_file:
                self._write_csv(temp_file, activities, self._get_activity_headers())
                return temp_path
        elif format_type == "json":
            with open(temp_path, 'w') as temp_file:
                json.dump(activities, temp_file, indent=2, default=str)
                return temp_path
        elif format_type == "pdf":
            return self._write_pdf(activities, self._get_activity_headers(), "activities", temp_path)
        else:
            # Default to JSON for other formats
            with open(temp_path, 'w') as temp_file:
                json.dump(activities, temp_file, indent=2, default=str)
                return temp_path
    
    def export_all_data(self, format_type: str, filters: Dict[str, Any]) -> str:
        """Export all data
        
        Args:
            format_type: Format to export in
            filters: Dictionary of filters to apply
            
        Returns:
            str: Path to the exported file
        """
        # Create temporary file
        all_data = {
            "employees": self._get_employee_data(filters),
            "payroll": self._get_payroll_data(filters),
            "overtime": self._get_overtime_data(filters),
            "activities": self._get_activity_data(filters)
        }
        
        if format_type == "excel":
            # For Excel, create a multi-sheet workbook
            return self._write_multi_sheet_excel(all_data)
        elif format_type == "csv":
            # For CSV, create a separate file for each data type and zip them
            temp_files = []
            try:
                with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".zip") as zip_file:
                    with zipfile.ZipFile(zip_file.name, 'w') as zipf:
                        for data_type, data in all_data.items():
                            csv_filename = f"{data_type}.csv"
                            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".csv") as csv_file:
                                self._write_csv(csv_file, data, self._get_headers_for_type(data_type))
                                zipf.write(csv_file.name, arcname=csv_filename)
                                temp_files.append(csv_file.name)
                    
                    return zip_file.name
            finally:
                # Clean up temporary CSV files
                for temp_file in temp_files:
                    try:
                        os.unlink(temp_file)
                    except OSError:
                        pass  # File might already be deleted
        elif format_type == "zip":
            # For ZIP, create a zip file containing all data types
            temp_files = []
            try:
                with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".zip") as zip_file:
                    with zipfile.ZipFile(zip_file.name, 'w') as zipf:
                        for data_type, data in all_data.items():
                            if data_type == "all":
                                continue  # Skip 'all' data type as it contains everything
                            
                            csv_filename = f"{data_type}.csv"
                            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".csv") as csv_file:
                                self._write_csv(csv_file, data, self._get_headers_for_type(data_type))
                                zipf.write(csv_file.name, arcname=csv_filename)
                                temp_files.append(csv_file.name)
                    
                    return zip_file.name
            finally:
                # Clean up temporary CSV files
                for temp_file in temp_files:
                    try:
                        os.unlink(temp_file)
                    except OSError:
                        pass  # File might already be deleted
        elif format_type == "json":
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".json") as temp_file:
                json.dump(all_data, temp_file, indent=2, default=str)
                return temp_file.name
        elif format_type == "pdf":
            # For PDF, create a combined PDF with all data
            return self._write_multi_section_pdf(all_data)
        else:
            # Default to JSON for other formats
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".json") as temp_file:
                json.dump(all_data, temp_file, indent=2, default=str)
                return temp_file.name
    
    def cleanup_old_exports(self, days_old: int) -> int:
        """Clean up old export files
        
        Args:
            days_old: Number of days old files should be to be cleaned up
            
        Returns:
            int: Number of files cleaned up
        """
        import glob
        
        cleaned_count = 0
        cutoff_date = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
        
        # Clean up files from all export directories
        for export_dir in self.export_dirs.values():
            if os.path.exists(export_dir):
                # Find all export files in this directory
                export_files = glob.glob(os.path.join(export_dir, "*.csv")) + \
                              glob.glob(os.path.join(export_dir, "*.json")) + \
                              glob.glob(os.path.join(export_dir, "*.xlsx")) + \
                              glob.glob(os.path.join(export_dir, "*.pdf")) + \
                              glob.glob(os.path.join(export_dir, "*.zip"))
                
                for file_path in export_files:
                    try:
                        file_mtime = os.path.getmtime(file_path)
                        if file_mtime < cutoff_date:
                            os.unlink(file_path)
                            cleaned_count += 1
                    except OSError:
                        # File might be in use or already deleted
                        pass
        
        return cleaned_count
    
    def _get_employee_data(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get employee data with filters
        
        Args:
            filters: Dictionary of filters to apply
            
        Returns:
            List[Dict[str, Any]]: List of employee records
        """
        query = self.db.query(User)
        
        if 'department_id' in filters:
            query = query.filter(User.department_id == filters['department_id'])
        
        if 'role_id' in filters:
            query = query.filter(User.role_id == filters['role_id'])
        
        if 'status' in filters:
            query = query.filter(User.status == filters['status'])
        
        employees = query.all()
        
        result = []
        for emp in employees:
            emp_dict = {
                "employee_id": emp.user_id,
                "first_name": emp.first_name,
                "last_name": emp.last_name,
                "email": emp.email,
                "phone": emp.phone_number,
                "role_id": emp.role_id,
                "role_name": emp.role_name,
                "department_id": emp.department_id,
                "department_name": emp.department.department_name if emp.department else None,
                "hourly_rate": float(emp.hourly_rate) if emp.hourly_rate else 0.0,
                "date_hired": emp.date_hired.isoformat() if emp.date_hired else None,
                "status": emp.status
            }
            result.append(emp_dict)
        
        return result
    
    def _clean_data_for_export(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Clean data for export by removing None values and empty strings
        
        Args:
            data: List of dictionaries containing data
            
        Returns:
            List[Dict[str, Any]]: Cleaned data
        """
        cleaned_data = []
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                if value is not None and value != '':
                    # Convert datetime objects to strings
                    if isinstance(value, (datetime, date)):
                        cleaned_row[key] = value.isoformat()
                    else:
                        cleaned_row[key] = value
                else:
                    # Convert None to empty string
                    cleaned_row[key] = ""
            cleaned_data.append(cleaned_row)
        return cleaned_data
    
    def _generate_filename(self, data_type: str, format_type: str, filters: Optional[Dict[str, Any]] = None) -> str:
        """Generate filename for export
        
        Args:
            data_type: Type of data being exported
            format_type: Format of the export
            filters: Optional filters to include in filename
            
        Returns:
            str: Generated filename
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{data_type}_export_{timestamp}.{format_type}"
        
        # Add filters to filename if provided
        if filters:
            filter_parts = []
            if 'start_date' in filters:
                start_date = filters['start_date']
                if isinstance(start_date, date):
                    start_date_str = start_date.strftime('%Y-%m-%d')
                else:
                    start_date_str = str(start_date)
                filter_parts.append(f"from_{start_date_str}")
            
            if 'end_date' in filters:
                end_date = filters['end_date']
                if isinstance(end_date, date):
                    end_date_str = end_date.strftime('%Y-%m-%d')
                else:
                    end_date_str = str(end_date)
                filter_parts.append(f"to_{end_date_str}")
            
            if 'department_id' in filters:
                filter_parts.append(f"dept_{filters['department_id']}")
            
            if filter_parts:
                filename = f"{data_type}_{'_'.join(filter_parts)}_{timestamp}.{format_type}"
        
        return filename
    
    def _get_payroll_data(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get payroll data with filters
        
        Args:
            filters: Dictionary of filters to apply
            
        Returns:
            List[Dict[str, Any]]: List of payroll records
        """
        query = self.db.query(Payroll)
        
        if 'start_date' in filters:
            query = query.filter(Payroll.cutoff_start >= filters['start_date'])
        
        if 'end_date' in filters:
            query = query.filter(Payroll.cutoff_end <= filters['end_date'])
        
        if 'department_id' in filters:
            query = query.join(User).filter(User.department_id == filters['department_id'])
        
        if 'user_id' in filters:
            query = query.join(User).filter(User.user_id == filters['user_id'])
        
        payrolls = query.all()
        
        result = []
        for payroll in payrolls:
            payroll_dict = {
                "payroll_id": payroll.payroll_id,
                "employee_id": payroll.user_id,
                "employee_name": f"{payroll.user.first_name} {payroll.user.last_name}",
                "pay_date": payroll.cutoff_end.isoformat() if payroll.cutoff_end else None,
                "basic_salary": float(payroll.basic_pay) if payroll.basic_pay else 0.0,
                "overtime_pay": float(payroll.overtime_pay) if payroll.overtime_pay else 0.0,
                "deductions": float(payroll.deductions) if payroll.deductions else 0.0,
                "net_salary": float(payroll.net_pay) if payroll.net_pay else 0.0,
                "created_at": payroll.generated_at.isoformat() if payroll.generated_at else None
            }
            result.append(payroll_dict)
        
        return result
    
    def _get_overtime_data(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get overtime data with filters
        
        Args:
            filters: Dictionary of filters to apply
            
        Returns:
            List[Dict[str, Any]]: List of overtime records
        """
        query = self.db.query(OvertimeRequest)
        
        if 'start_date' in filters:
            query = query.filter(OvertimeRequest.date >= filters['start_date'])
        
        if 'end_date' in filters:
            query = query.filter(OvertimeRequest.date <= filters['end_date'])
        
        if 'status' in filters:
            query = query.filter(OvertimeRequest.status == filters['status'])
        
        if 'user_id' in filters:
            query = query.filter(OvertimeRequest.user_id == filters['user_id'])
        
        overtime = query.all()
        
        result = []
        for ot in overtime:
            ot_dict = {
                "overtime_id": ot.ot_id,
                "employee_id": ot.user_id,
                "employee_name": f"{ot.user.first_name} {ot.user.last_name}",
                "overtime_date": ot.date.isoformat() if ot.date else None,
                "hours_worked": float(ot.hours_requested) if ot.hours_requested else 0.0,
                "rate_per_hour": 0.0,  # Not available in model
                "overtime_pay": 0.0,  # Not available in model
                "status": ot.status,
                "created_at": ot.approved_at.isoformat() if ot.approved_at else None,
                "updated_at": ot.approved_at.isoformat() if ot.approved_at else None
            }
            result.append(ot_dict)
        
        return result
    
    def _get_activity_data(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get activity data with filters
        
        Args:
            filters: Dictionary of filters to apply
            
        Returns:
            List[Dict[str, Any]]: List of activity records
        """
        query = self.db.query(ActivityLog)
        
        if 'start_date' in filters:
            query = query.filter(ActivityLog.timestamp >= filters['start_date'])
        
        if 'end_date' in filters:
            query = query.filter(ActivityLog.timestamp <= filters['end_date'])
        
        if 'user_id' in filters:
            query = query.filter(ActivityLog.user_id == filters['user_id'])
        
        if 'action' in filters:
            query = query.filter(ActivityLog.action.like(f"%{filters['action']}%"))
        
        activities = query.all()
        
        result = []
        for act in activities:
            act_dict = {
                "activity_id": act.log_id,
                "employee_id": act.user_id,
                "employee_name": f"{act.user.first_name} {act.user.last_name}",
                "activity_date": act.timestamp.isoformat() if act.timestamp else None,
                "action": act.action,
                "details": act.details,
                "created_at": act.timestamp.isoformat() if act.timestamp else None,
                "updated_at": act.timestamp.isoformat() if act.timestamp else None
            }
            result.append(act_dict)
        
        return result
    
    def _write_csv(self, file, data: List[Dict[str, Any]], headers: List[str]):
        """Write data to CSV file
        
        Args:
            file: File object to write to
            data: List of dictionaries containing data
            headers: List of column headers
        """
        if not data:
            # Create empty CSV with headers only
            writer = csv.DictWriter(file, fieldnames=headers)
            writer.writeheader()
            return
        
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        
        for row in data:
            # Convert datetime objects to strings for CSV
            csv_row = {}
            for key, value in row.items():
                if isinstance(value, (datetime, date)):
                    csv_row[key] = value.isoformat()
                elif value is None:
                    csv_row[key] = ""
                else:
                    csv_row[key] = value
            
            writer.writerow(csv_row)
    
    def _get_employee_headers(self) -> List[str]:
        """Get employee data headers for CSV export
        
        Returns:
            List[str]: List of employee data headers
        """
        return [
            "employee_id", "first_name", "last_name", "email", "phone",
            "role_id", "role_name", "department_id", "department_name",
            "hourly_rate", "date_hired", "status", "created_at", "updated_at"
        ]
    
    def _get_payroll_headers(self) -> List[str]:
        """Get payroll data headers for CSV export
        
        Returns:
            List[str]: List of payroll data headers
        """
        return [
            "payroll_id", "employee_id", "employee_name", "pay_date",
            "basic_salary", "overtime_pay", "deductions", "net_salary",
            "payroll_status", "created_at"
        ]
    
    def _get_overtime_headers(self) -> List[str]:
        """Get overtime data headers for CSV export
        
        Returns:
            List[str]: List of overtime data headers
        """
        return [
            "overtime_id", "employee_id", "employee_name", "overtime_date",
            "hours_worked", "rate_per_hour", "overtime_pay", "status",
            "created_at", "updated_at"
        ]
    
    def _get_activity_headers(self) -> List[str]:
        """Get activity data headers for CSV export
        
        Returns:
            List[str]: List of activity data headers
        """
        return [
            "activity_id", "employee_id", "employee_name", "activity_date",
            "action", "details", "created_at", "updated_at"
        ]
    
    def _get_headers_for_type(self, data_type: str) -> List[str]:
        """Get headers for specific data type
        
        Args:
            data_type: Type of data to get headers for
            
        Returns:
            List[str]: List of headers for the specified data type
        """
        if data_type == "employees":
            return self._get_employee_headers()
        elif data_type == "payroll":
            return self._get_payroll_headers()
        elif data_type == "overtime":
            return self._get_overtime_headers()
        elif data_type == "activities":
            return self._get_activity_headers()
        else:
            return []
    
    def _write_excel(self, data: List[Dict[str, Any]], headers: List[str], sheet_name: str, file_path: str = None) -> str:
        """Write data to Excel file
        
        Args:
            data: List of dictionaries containing data
            headers: List of column headers
            sheet_name: Name of the worksheet
            file_path: Optional path to save the file
            
        Returns:
            str: Path to the Excel file
        """
        if file_path:
            temp_file = open(file_path, 'wb')
        else:
            if not data:
                # No data to write, create empty file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    worksheet.title = sheet_name
                    
                    # Add headers
                    for col_num, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    workbook.save(temp_file.name)
                    return temp_file.name
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = sheet_name
                
                # Add headers
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data
                for row_num, row_data in enumerate(data, 2):
                    for col_num, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        value = row_data.get(header, '')
                        
                        # Handle datetime objects
                        if isinstance(value, (datetime, date)):
                            cell.value = value.isoformat()
                        else:
                            cell.value = value
                
                # Auto-adjust column widths
                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[column_letter].width = 15
                
                workbook.save(temp_file.name)
                return temp_file.name
        
        # Handle the case where file_path is provided
        with open(file_path, 'wb') as f:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    value = row_data.get(header, '')
                    
                    # Handle datetime objects
                    if isinstance(value, (datetime, date)):
                        cell.value = value.isoformat()
                    else:
                        cell.value = value
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 15
            
            workbook.save(file_path)
            return file_path
    
    def _write_multi_sheet_excel(self, all_data: Dict[str, List[Dict[str, Any]]]) -> str:
        """Write data to multi-sheet Excel file
        
        Args:
            all_data: Dictionary containing data for different types
            
        Returns:
            str: Path to the Excel file
        """
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Add data sheets
            for sheet_name, data in all_data.items():
                if data:  # Only add sheets with data
                    worksheet = workbook.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                    
                    # Get headers for this data type
                    headers = self._get_headers_for_type(sheet_name)
                    
                    # Add headers
                    for col_num, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Add data
                    for row_num, row_data in enumerate(data, 2):
                        for col_num, header in enumerate(headers, 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            value = row_data.get(header, '')
                            
                            # Handle datetime objects
                            if isinstance(value, (datetime, date)):
                                cell.value = value.isoformat()
                            else:
                                cell.value = value
                    
                    # Auto-adjust column widths
                    for col_num in range(1, len(headers) + 1):
                        column_letter = get_column_letter(col_num)
                        worksheet.column_dimensions[column_letter].width = 15
            
            workbook.save(temp_file.name)
            return temp_file.name
    
    def _write_pdf(self, data: List[Dict[str, Any]], headers: List[str], data_type: str, file_path: str = None) -> str:
        """Write data to PDF file
        
        Args:
            data: List of dictionaries containing data
            headers: List of column headers
            data_type: Type of data being exported
            file_path: Optional path to save the file
            
        Returns:
            str: Path to the PDF file
        """
        from backend.services.pdf_generator import PDFGeneratorService
        
        if file_path:
            pdf_generator = PDFGeneratorService()
            pdf_generator.generate_data_export(file_path, data, headers, data_type)
            return file_path
        else:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                pdf_generator = PDFGeneratorService()
                pdf_generator.generate_data_export(temp_file.name, data, headers, data_type)
                return temp_file.name
    
    def _write_multi_section_pdf(self, all_data: Dict[str, List[Dict[str, Any]]]) -> str:
        """Write data to multi-section PDF file
        
        Args:
            all_data: Dictionary containing data for different types
            
        Returns:
            str: Path to the PDF file
        """
        from backend.services.pdf_generator import PDFGeneratorService
        
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_generator = PDFGeneratorService()
            
            # Create sections for each data type
            for data_type, data in all_data.items():
                if data:  # Only add sections with data
                    headers = self._get_headers_for_type(data_type)
                    pdf_generator.add_data_section(temp_file.name, data, headers, data_type)
            
            return temp_file.name